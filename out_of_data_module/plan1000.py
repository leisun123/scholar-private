#coding:utf-8

import sys
import importlib
importlib.reload(sys)
import os
import openpyxl

from pdfminer.pdfparser import PDFParser,PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LTTextBoxHorizontal,LAParams
from pdfminer.pdfinterp import PDFTextExtractionNotAllowed

path = r'/Users/sunlei/Downloads/国家千人计划名单(5批).pdf'
# os.remove('./1.txt')
def parse():
    fp = open(path, 'rb') # 以二进制读模式打开
    #用文件对象来创建一个pdf文档分析器
    praser = PDFParser(fp)
    # 创建一个PDF文档
    doc = PDFDocument()
    # 连接分析器 与文档对象
    praser.set_document(doc)
    doc.set_parser(praser)

    # 提供初始化密码
    # 如果没有密码 就创建一个空的字符串
    doc.initialize()

    # 检测文档是否提供txt转换，不提供就忽略
    if not doc.is_extractable:
        raise PDFTextExtractionNotAllowed
    else:
        # 创建PDf 资源管理器 来管理共享资源
        rsrcmgr = PDFResourceManager()
        # 创建一个PDF设备对象
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        # 创建一个PDF解释器对象
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        # 循环遍历列表，每次处理一个page的内容
        for page in doc.get_pages():  # doc.get_pages() 获取page列表
            interpreter.process_page(page)
            # 接受该页面的LTPage对象
            layout = device.get_result()
            # 这里layout是一个LTPage对象 里面存放着 这个page解析出的各种对象 一般包括LTTextBox, LTFigure, LTImage, LTTextBoxHorizontal 等等 想要获取文本就获得对象的text属性，
            for x in layout:
                if (isinstance(x, LTTextBoxHorizontal)):
                    with open(r'./1.txt', 'a') as f:
                        results = x.get_text()
                        print(results)
                        f.write(results + '\n')

def getdic():
    i = 1
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'plan1000'
    lis = ['姓名','职称','单位']
    for i in range(0,3):
        sheet.cell(column=i+1,row=1,value=str(lis[i]))
    a = []
    with open ("./name.txt","r") as f:
        line = f.readline()
        while line:
            if ":" not in line and line != '\n':
                res = line.replace(' \n','')
                if res != "":
                    res = res
            elif line != '\n':
                tup = (line.split(':')[0],line.split(':')[-1].replace(' \n',''),res)
                for j in range(0,3):
                    sheet.cell(column=j+1,row=i,value=str(tup[j]))
                i = i + 1
                print(tup)
            line = f.readline()
        wb.save('./plan1000.xlsx')
        f.close()
        print(a)
if __name__ == '__main__':
    # parse()
    getdic()